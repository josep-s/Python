import openpyxl
wb = openpyxl.load_workbook("D:\\t129627\\Documents\\TablaAccesosFuncionaliades.xlsx")
ws = wb["Listado"] # També es podria accedir a la fulla del llibre amb ws = wb.active
print("Nombre de files: " + str(ws.max_row) + "\nNombre de columnes: " + str(ws.max_column) + "\n\n")
print("El valor de la cela A1 és: " + ws['A1'].value + "\n\n")
Capçalera = [ws.cell(row=1,column=i).value for i in range(1,ws.max_column+1)]
print("Capçalera:")
print(Capçalera)
print("\n\n")
print("10 valors de la columna CENTRO:")
Dades=[ws.cell(row=i,column=2).value for i in range(2,12)]
print(Dades)
print("\n\n")
print("10 primeres files:")
# reading data from a range of cells (from column 1 to 6)

my_list = list()

for value in ws.iter_rows(
    min_row=1, max_row=11, min_col=1, max_col=6, 
    values_only=True):
    my_list.append(value)
    
for ele1,ele2,ele3,ele4,ele5,ele6 in my_list:
    (print ("{:<8}{:<35}{:<10}{:<10}{:<15}{:<15}".format(ele1,ele2,ele3,ele4,ele5,ele6)))
# Tot açò está tret de la pàgina https://www.datacamp.com/es/tutorial/python-excel-tutorial. N'hi ha molta més informació.